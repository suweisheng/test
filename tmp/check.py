# encoding: utf8

from concurrent.futures import ThreadPoolExecutor
import socket
import uuid
import requests
import json
import re
import os
import time
from datetime import datetime
import xlrd
import xlsxwriter
import shutil
import pymongo
import yaml
import multiprocessing
import openpyxl
import xlrd

def run(func):
    start_time = time.time()
    func()
    end_time = time.time()
    print("\n[Finish {:.4f} s]".format(end_time-start_time))

def time_second():
    return int(time.time())

def data_to_ts(time_str):
    date_object = datetime.strptime(time_str, r'%Y-%m-%d %H:%M:%S')
    return int(time.mktime(date_object.timetuple()))

def format_time(ts):
    return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(ts))

def format_day(ts):
    return time.strftime("%Y-%m-%d", time.localtime(ts))

def read_xlsx_file(filename, sheet_index=0, sheet_name=None):
    workbook = openpyxl.load_workbook(filename.decode("utf8"))
    sheet = None
    if sheet_name:
        sheet = workbook.sheet_by_name(sheet_name)
    else:
        sheet = workbook.worksheets[sheet_index]
    lines = []
    for row in sheet.iter_rows(values_only=True):
        lines.append(row)
    workbook.close()
    return lines

def read_xls_file(filename, sheet_index=0, sheet_name=None):
    workbook = xlrd.open_workbook(filename.decode("utf8"))
    sheet = None
    if sheet_name:
        sheet = workbook.sheet
    else:
        sheet = workbook.sheet_by_index(0)
    lines = []
    for row_idx in range(0, sheet.nrows):
        row_data = []
        for col_idx in range(0, sheet.ncols):
            cell_value = sheet.cell_value(row_idx, col_idx)
            cell_type = sheet.cell_type(row_idx, col_idx)
            if cell_type == xlrd.XL_CELL_DATE:
                year, month, day, hour, minute, second = xlrd.xldate_as_tuple(cell_value, workbook.datemode)
                cell_value = int(time.mktime([year, month, day, hour, minute, second, 0, 0, 0]))
            row_data.append(cell_value)
        lines.append(row_data)

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 

def gethostbyaddr(ip_address):
    ret = ""
    try:
        ret = socket.gethostbyaddr(ip_address)[0]
    except socket.herror as err:
        pass
    return  ret

def checklocalnet():
    ip_list = []
    for i in range(1, 255):
        ip_list.append("192.168.170.{}".format(i))
    with ThreadPoolExecutor(max_workers=100) as pool:
        results = list(pool.map(gethostbyaddr, ip_list))

    ok_str, err_str = "", ""
    for i in range(len(ip_list)):
        if results[i]:
            ok_str = ok_str + "ip:{}, host:{}\n".format(ip_list[i], results[i])
        else:
            err_str = err_str + ip_list[i] + "\n"

    with open("checkIp.log", "w") as f:
        f.write(ok_str)
        f.write("==========> \n")
        f.write(err_str)

# checklocalnet()

def find_files_with_keyword(path_list, keyword):
    keyword = keyword.lower()
    found_ret = []
    for path in path_list:
        for foldername, subfolders, filenames in os.walk(path):
            for foldname in subfolders:
                if keyword in foldname.lower():  # 忽略大小写
                    found_ret.append(os.path.join(foldername, foldname))
            for filename in filenames:
                if keyword in filename.lower():  # 忽略大小写
                    found_ret.append(os.path.join(foldername, filename))
    return found_ret

# ret = find_files_with_keyword(["C:\\Users\\Admin", "C:\\Program Files", "C:\\ProgramData"], "baiduyun")
# for x in ret: print(x)

def fix_flower_act():
    record_path = "C:\\Users\\Admin\Desktop\\record"
    all_record = []
    for file in os.listdir(record_path):
        file_path = os.path.join(record_path, file)
        with open(file_path, "r") as f:
            for line in f:
                info = line.split("; ")
                info[-1] = info[-1].strip()
                info[0] = data_to_ts(info[0])
                all_record.append(info)

    act_path = "C:\\Users\\Admin\Desktop\\act"
    all_act_dict = {}
    for file in os.listdir(act_path):
        file_path = os.path.join(act_path, file)
        with open(file_path, "r") as f:
            for line in f:
                info = line.strip().split("; ")
                if all_act_dict.has_key(info[-1]):
                    raise Exception(u"duplicate uuid:"+info[-1])
                else:
                    all_act_dict[info[-1]] = data_to_ts(info[0])

    server2record = {}
    for v in all_record:
        reveive = v[2]
        if not all_act_dict.has_key(reveive) or all_act_dict[reveive] > v[0]:
            server_id = int(v[1][0:5])
            if server2record.has_key(server_id):
                server2record[server_id].append(v)
            else:
                server2record[server_id] = [v]

    f1 = open("flower.log", "w")
    f2 = open("flower_server.log", "w")
    server_id_list = server2record.keys()
    server_id_list.sort()
    for server_id in server_id_list:
        f2.write("==================================================================================> [%d] - %d\n" % (server_id, len(server2record[server_id])))
        for v in server2record[server_id]:
            content = str("{{ts={}, send='{}', reveive='{}', num={}}},\n".format(v[0], v[1], v[2], int(v[3])))
            f1.write(content)
            f2.write(content)
    f1.close()
    f2.close()


def fix_follow_rank_mail():
    zone_info_list = [
        # game_id_min, game_id_max, vserver_id_begin, zone_id_begin, merged_zone_id_max
        [20001, 20099, 21001, 1, 24],
        [27001, 27099, 27001, 27001, 27004],
        [30000, 30099, 30001, 30001, 30042],
        [33000, 33099, 33001, 33001, 33002],
        [33100, 33199, 33101, 33101, 0],
        [36000, 36099, 36001, 36001, 0],
        [40000, 40099, 40001, 40001, 40030],
        [46000, 46099, 46001, 46001, 46036],
    ]
    # 猎码：20000-20099
    # 来玩：27000-27099
    # 微信：30000-30099
    # 圣本：33000-33099
    # 圣本2: 33100-33199
    # 抖音：36000-36099
    # 漫灵：40000-40099
    # 漫灵买量：46000-46099
    def merge_id(x):
        if x%2 == 0:
            return x - 1
        else:
            return x

    def parse_uuid(uuid):
        uuid = str(uuid)
        game_id = int(uuid[0:5])
        vserver_id = int(uuid[5:10])
        zone_id = 0
        for info in zone_info_list:
            if game_id >= info[0] and game_id <= info[1]:
                vserver_id_begin = info[2]
                zone_id_begin = info[3]
                merged_zone_id_max = info[4]

                zone_id = (vserver_id - vserver_id_begin) / 4 + zone_id_begin
                zone_id = int(zone_id)
                if zone_id <= merged_zone_id_max:
                    zone_id = merge_id(zone_id)
                return info[0], game_id, zone_id
        # assert(False, "invalid uuid" + uuid)

    record_path = "C:\\Users\\Admin\Desktop\\offlinemail"
    record_list = []
    for file in os.listdir(record_path):
        with open(os.path.join(record_path, file), "r") as f:
            for line in f:
                info = line.strip().split("; ")
                group_code, game_id, zone_id = parse_uuid(info[3])
                record_list.append(dict(
                    ts = data_to_ts(info[0]),
                    mail_guid = info[1],
                    mail_id = info[2],
                    uuid = info[3],
                    rank = int(info[4]),
                    group_code = group_code,
                    zone_id = zone_id,
                ))

    zone2rank = {}
    group2rank = {}
    error_list = []
    for v in record_list:
        mail_id = v["mail_id"]
        rank = v["rank"]
        zone_id = v["zone_id"]
        group_code = v["group_code"]

        rank_data = None
        if mail_id == "114" or mail_id == "116":
            if zone_id not in zone2rank:
                zone2rank[zone_id] = {}
            rank_data = zone2rank[zone_id]
        elif mail_id == "113" or mail_id == "115":
            if group_code not in group2rank:
                group2rank[group_code] = {}
            rank_data = group2rank[group_code]
        else:
            raise Exception("=== error mail_id:%s" % str(mail_id))
        
        if mail_id not in rank_data:
            rank_data[mail_id] = {}
    
        rank_list = rank_data[mail_id]

        last_info = rank_list.get(rank)
        if last_info:
            if v["ts"] < last_info["ts"]:
                rank_list[rank] = v
                error_list.append(last_info)
            else:
                error_list.append(v)
        else:
            rank_list[rank] = v


    zone_list = []
    for _, data in zone2rank.items():
        for _, v in data.items():
            zone_list.extend(v.values())

    group_list = []
    for _, data in group2rank.items():
        for _, v in data.items():
            group_list.extend(v.values())

    error_list.sort(key=lambda x: (x['mail_id'], x['rank']))
    
    print(len(record_list), len(zone_list), len(group_list), len(error_list),
            len(zone_list)+len(group_list)+len(error_list))


    with open("zone_rank.log", "w") as f:
        for v in zone_list:
            content = "{{zone_id={}, mail_id={}, rank={}, uuid='{}', ts={}, mail_guid='{}'}},\n".format(
                            v['zone_id'], v['mail_id'], v['rank'], v['uuid'], v["ts"], v['mail_guid'])
            f.write(content)

    with open("group_rank.log", "w") as f:
        for v in group_list:
            content = "{{group_id={}, mail_id={}, rank={}, uuid='{}', ts={}, mail_guid='{}'}},\n".format(
                            v['group_code'], v['mail_id'], v['rank'], v['uuid'], v['ts'], v['mail_guid'])
            f.write(content)

    with open("error_rank.log", "w") as f:
        for v in error_list:
            content = "{{mail_id={}, rank={}, uuid='{}', ts={}, mail_guid='{}'}},\n".format(
                            v['mail_id'], v['rank'], v['uuid'], v['ts'], v['mail_guid'])
            f.write(content)

    uuid = "460114614000768"
    for v in zone_list:
        if v["uuid"] == uuid:
            print("zone:", v)
    for v in group_list:
        if v["uuid"] == uuid:
            print("group:", v)
    for v in error_list:
        if v["uuid"] == uuid:
            print("error:", v)

# fix_follow_rank_mail()


def format_thinking(input_file, output_file):
    content = None
    with open(input_file, 'r') as f_in:
        content = f_in.read()
    content = content.replace(" \n", "") # 原先成功换行的
    content = content.replace("\n\n", "") # 去掉空行

    lines = content.split("* ")[1:]

    num = 80
    with open(output_file, 'w') as f_out:
        for line in lines:
            s_bytes = line.decode('utf-8')
            real_str = ""
            for j in range(0, len(s_bytes), num):
                if j != 0:
                    real_str = real_str + ' \n'
                real_str = real_str + s_bytes[j:j+num]
            real_str = "* " + real_str.encode('utf8') + "\n\n"
            f_out.write(real_str)

# format_thinking('D:\\LearnCode\\Note\\thinking.md', 'D:\\LearnCode\\Note\\thinking.md')

def guo_kao_filter(data, filter_dict, must_dict):
    result = []
    new_must_dict = {}
    for biao_tou, key_list in must_dict.items():
        index = data["biao_tou_list"].index(biao_tou)
        new_must_dict[index] = key_list
    new_filter_dict = {}
    for biao_tou, key_list in filter_dict.items():
        index = data["biao_tou_list"].index(biao_tou)
        new_filter_dict[index] = key_list

    for line in data["data_list"]:
        is_ok = False
        for biao_tou, key_list in new_must_dict.items():
            content = line[biao_tou]
            for key in key_list:
                if key in content:
                    is_ok = True
                    break
            if is_ok == True:
                break
        
        if is_ok == True:
            for biao_tou, key_list in new_filter_dict.items():
                content = line[biao_tou]
                for key in key_list:
                    if key in content:
                        is_ok = False
                        break
                if is_ok == False:
                    break

        if is_ok == True:
            result.append(line)

    return result

def guo_kao_xuan_gang():
    filename = "D:\\LearnCode\\中央机关及其直属机构2024年度考试录用公务员招考简章.xls"
    workbook = xlrd.open_workbook(filename.decode("utf8"))
    sheet_list = workbook.sheets()
    all_data = []
    for sheet in sheet_list:
        biao_tou_list = sheet.row_values(rowx=1) # 表头,第二行
        biao_tou_len = len(biao_tou_list)

        data_list = []
        for r in range(2, sheet.nrows):
            data_list.append(sheet.row_values(rowx=r, start_colx=0, end_colx=biao_tou_len))

        all_data.append({
            "name" : sheet.name,
            "biao_tou_list" : biao_tou_list,
            "data_list" : data_list
        })
    
    # 过滤
    filter = {
        u'政治面貌' : [u"中共党员", u"中共党员或共青团员"],
        u'学历' : [u"仅限硕士研究生", u"仅限博士研究生", u"硕士研究生及以上"],
        u'学位' : [u"硕士", u"博士",],
        u'服务基层项目工作经历' : [u"大学生村官", u"三支一扶", u"大学生志愿服务西部计划"],
        u'备注' : [u"限应届毕业生", u"应届毕业生", u"2024届高校毕业生", u"应届高校毕业生", u"大学英语6级", u"英语六级",
                    u"普通话等级达到", u"国有企事业单位在职人员", u"雅思成绩", u"专业四级", u"专业八级", u"女性",
                    u"注册税务师证书", u"PMP", u"税务系统", u"计算机二级", u"律师从业资格", u"在北京税务系统", u"法律"
                    u"限云南省户籍", u"限应届", u"商业性金融机构", u"在办税服务厅工作", u"有参与课题研究经验",
                    u"铁路运输",
                ],
    }
    must = {
        u'专业' : [u"工业工程"],
    }

    for data in all_data:
        ret = guo_kao_filter(data, filter, must)
        data["filter_ret"] = ret
        print(len(data["data_list"]), len(data["filter_ret"]))

    with open("gwy_heshi.log", "w") as f:
        for data in all_data:
            for line in data["filter_ret"]:
                f.write(str(line).decode('unicode_escape').encode("utf8") + "\n")
                # f.write(line[22].encode("utf8") + "\n")

    with xlsxwriter.Workbook("gwy_heshi.xlsx") as workbook:
        for data in all_data:
            worksheet = workbook.add_worksheet(data["name"])
            worksheet.write_row(0, 0, data["biao_tou_list"])
            for i in range(1, len(data["filter_ret"])):
                worksheet.write_row(i, 0, data["filter_ret"][i-1])

# guo_kao_xuan_gang()

def nan_san_gou_che():
    filename = "E:\\temp\\爱车南山购第三轮汽车专项促销费活动拟发放名单公示.xls"
    workbook = xlrd.open_workbook(filename.decode("utf8"))
    all_data = []
    sheet_list = workbook.sheets()
    for sheet in sheet_list:
        if sheet.cell_value(rowx=0, colx=0) == "":
            continue
        biao_tou_list = sheet.row_values(rowx=1) # 表头,第二行
        biao_tou_len = len(biao_tou_list)

        data_list = []
        for r in range(2, sheet.nrows):
            data_list.append(sheet.row_values(rowx=r, start_colx=0, end_colx=biao_tou_len))

        all_data.append({
            "name" : sheet.name,
            "biao_tou_list" : biao_tou_list,
            "data_list" : data_list
        })

    shen_fen_zheng_dict = dict()
    total_amount = 0
    for data in all_data:
        amount = 0
        for line in data["data_list"]:
            key = line[2] + "_" + line[3]
            if shen_fen_zheng_dict.has_key(key):
                shen_fen_zheng_dict[key] = shen_fen_zheng_dict[key] + 1
            else:
                shen_fen_zheng_dict[key] = 1
            amount += line[5]
            print(amount)
        total_amount += amount
    for key, num in shen_fen_zheng_dict.items():
        if num > 1:
            print(key, num)
    print(100000000-total_amount)

# nan_san_gou_che()

def xiaomi_guoji():
    import base64
    import hashlib
    import hmac

    app_id = "1111"
    app_secret = "222323fdsfafa"

    purchase_token = "PCN220513171445751001000014572350"
    package_name = "com.kachishop"
    product_id = "Game-1"

    path = "/{}/developer/v1/applications/{}/purchases/products/{}/tokens/{}".format(purchase_token[1:3], package_name, product_id, purchase_token)
    rand_str = uuid.uuid1()
    rand_str = "sdfsdfafdsfsdfdsfsadfasdfasdfasdf"
    sign = "GET\n\napplication/json\n\nx-ias-sign-nonce:{}\n{}".format(rand_str, path)
    sign = hmac.new(app_secret, sign, hashlib.sha1)
    sign = str(base64.b64encode(sign.digest()))
    print(sign)

# xiaomi_guoji()

def fix_game_log():
    filename = "C:\\Users\\Admin\\Desktop\\fb_log1.xlsx"
    lines = read_xlsx_file(filename)

    data_dict = {}
    for line in lines[1:]:
        time = format_day(line[0]/1000)
        if time not in data_dict:
            data_dict[time] = {u"FB社群":set(), u"FB主页":set()}
        data_dict[time][line[3]].add(line[2])


    data_list = []
    for time, v in data_dict.items():
        data_list.append([time, len(v[u"FB社群"]), len(v[u"FB主页"])])

    with xlsxwriter.Workbook("fb_log.xlsx") as workbook:
        worksheet = workbook.add_worksheet("sheet1")
        num = 1
        for row in data_list:
            worksheet.write_row(num, 0, row)
            num += 1

# fix_game_log()

# filename = "C:\\Users\\Admin\\Desktop\\gamelog.xls"
# lines = read_xls_file(filename)

url = "http://minigame-api.skyblade88.com/cp/game-check"
data = dict(
    signature = '11',
    token='11',
    accountId ='126144',
    appId = 20007
)
cont = requests.post(url, data)
print(cont.json())
